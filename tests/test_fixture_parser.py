"""Fixture-based end-to-end tests for parse_s2d().

A synthetic S2D workbook is built programmatically at test-session start
using generic placeholder data (no real customer values).  parse_s2d()
is called once; all tests share the result via session fixtures.

Covers:
- Accinfo   → project_metadata, global_network_context, ignored_labels
- NWInfo    → network.subnets (group/env/purpose_key), network.routing_table
- ServerInfo → vmware_vm (phost split, compute, network, availability, metadata)
- FileSystem → filesystem (size anomaly: int vs formula string)
- SecurityGroup → firewall_rule_candidates (port normalization, booleans)
"""
import json
import pytest
import openpyxl

from src.parser import parse_s2d

# ---------------------------------------------------------------------------
# Synthetic workbook builder
# ---------------------------------------------------------------------------

def _build_accinfo(wb):
    ws = wb.create_sheet("Accinfo")

    def row(label=None, value=None):
        r = [None] * 17
        r[0] = label
        r[6] = value
        return r

    ws.append(row())                                              # blank
    ws.append(row("General Information"))                         # section header
    ws.append(row("Customer Name",                "Generic Corp"))
    ws.append(row("Customer (CID)",               "GCI"))
    ws.append(row())                                              # blank
    ws.append(row("IP Range",                     "10.0.0.0/22"))
    ws.append(row("DR IP Range",                  "10.4.0.0/22"))
    ws.append(row("Domain Name",                  "sap.generic.net"))
    ws.append(row("Customer Connectivity",         "Cloud Peering"))
    ws.append(row("Customer N/W",                  "10.100.0.0/24\n10.100.1.0/24"))
    ws.append(row("Customer DNS Server (Primary)", "10.0.0.10"))
    ws.append(row("Customer DNS Server (Secondary)", "10.0.0.11"))
    ws.append(row("System Timezone",               "KST (UTC+09:00)"))
    ws.append(row("VPC Name",                     None))          # out-of-scope → ignored
    ws.append(row("SID Information"))                             # defer sentinel


def _build_nwinfo(wb):
    ws = wb.create_sheet("NWInfo")
    N = 8  # generous column count

    def r(*vals):
        lst = list(vals)
        lst += [None] * (N - len(lst))
        return lst

    # Customer — PRD block
    ws.append(r("Customer"))
    ws.append(r(None, "PRD/non-PRD"))
    ws.append(r(None, "No.", "Purpose", "IP Range", "NAT IP Range", "Hostname"))
    ws.append(r(None, 1, "Admin",              "10.1.0.0/24",  None,              None))
    ws.append(r(None, 2, "Production Service", "10.0.0.0/24",  None,              "vhgeneric*"))
    ws.append(r())   # blank — must be skipped, NOT terminate
    ws.append(r(None, 3, "Backup",             "10.3.0.0/24",  None,              None))

    # Customer — DR block
    ws.append(r(None, "DR"))
    ws.append(r(None, "No.", "Purpose", "IP Range", "NAT IP Range", "Hostname"))
    ws.append(r(None, 1, "Admin",              "10.4.1.0/24",  None,              None))

    # Internal — PRD block
    ws.append(r("Internal - Storage Network"))
    ws.append(r(None, "PRD/non-PRD"))
    ws.append(r(None, "No.", "Purpose", "IP Range", "NAT IP Range", "Hostname"))
    ws.append(r(None, 1, "Storage",            "10.2.0.0/24",  None,              None))

    # Routing table
    ws.append(r("Server Routing Table"))
    ws.append(r("No.", "Routing Name", "Target", "Source", "Remark"))
    ws.append(r(1, "Customer Route", "10.100.0.0/16", "10.0.0.1", "primary"))
    ws.append(r(2, "DR Route",       "10.104.0.0/16", "10.0.0.1", None))


def _build_serverinfo(wb):
    ws = wb.create_sheet("ServerInfo")
    N = 43

    def blank_row():
        return [None] * N

    # Row 1: annotation
    ws.append(blank_row())

    # Row 2: column headers (partial)
    hdr = blank_row()
    hdr[0]  = "Phase";         hdr[1]  = "vhost";       hdr[2]  = "phost"
    hdr[3]  = "Landscape";     hdr[4]  = "SID";         hdr[5]  = "Type"
    hdr[6]  = "Main Solution"; hdr[9]  = "Type";        hdr[10] = "Role"
    hdr[13] = "CDD";           hdr[14] = "SCP Image";   hdr[15] = "OS Version"
    hdr[16] = "CPU";           hdr[17] = "Mem (GB)";    hdr[18] = "Appl (GB)"
    hdr[19] = "NFS";           hdr[20] = "SVC Host";    hdr[21] = "SVC IP"
    hdr[22] = "ADM Host";      hdr[23] = "ADM IP";      hdr[30] = "SLA"
    hdr[31] = "HA";            hdr[32] = "DR";          hdr[33] = "Pacemaker"
    ws.append(hdr)

    # Row 3: merged sub-header (blank)
    ws.append(blank_row())

    # Row 4: server entry (entry_type="server" → included)
    srv = blank_row()
    srv[0]  = 1
    srv[1]  = "vhgenericap01"
    srv[2]  = "phgenericap0110.1.0.11"   # phost + admin_ip concatenated
    srv[3]  = "PRD"
    srv[4]  = "TST"
    srv[5]  = "AP"
    srv[6]  = "S/4HANA"
    srv[7]  = "S/4HANA"
    srv[8]  = "APP"
    srv[9]  = "server"
    srv[10] = "PAS#01"
    srv[11] = "01"
    srv[12] = "VM"
    srv[14] = "RHEL9-GI"
    srv[15] = "RHEL 9.4"
    srv[16] = 8
    srv[17] = 32
    srv[18] = 100
    srv[19] = "N/A"         # nfs → None
    srv[20] = "vhgenericap01"
    srv[21] = "10.0.0.11"
    srv[22] = "phgenericap01"
    srv[23] = "10.1.0.11"   # admin_ip — also used as known_ip for phost split
    srv[24] = "192.168.0.11"
    srv[30] = "99.90%"
    srv[31] = "O"           # ha → True
    srv[32] = "X"           # dr → False
    srv[33] = "O"           # pacemaker → True
    srv[34] = "X"           # db_sr → False
    srv[36] = "O"           # backup_enabled → True
    ws.append(srv)

    # Row 5: vip entry (entry_type="vip" → excluded)
    vip = blank_row()
    vip[1] = "vhgenericvip01"
    vip[9] = "vip"
    ws.append(vip)


def _build_filesystem(wb):
    ws = wb.create_sheet("FileSystem")
    N = 19

    def blank_row():
        return [None] * N

    # Row 1: annotation
    ws.append(blank_row())

    # Row 2: column headers
    ws.append(["No.", "Landscape", "SID", "Server Type", "FS Cat1",
               "FS Cat2", "SLA Tier", "HA", "FS Count", "FS Seq",
               "Hostname", "Admin IP", "Mount Point", "Size (GB)", "FS Type",
               "VG Name", "NFS Group", "Remark", "Check"])

    # Row 3: plain-integer size row
    r1 = blank_row()
    r1[0]  = 1
    r1[1]  = "PRD"
    r1[2]  = "TST"
    r1[3]  = "AP"
    r1[4]  = "APP"
    r1[5]  = "data"
    r1[6]  = "Gold"
    r1[7]  = "O"    # ha_flag → True
    r1[8]  = 1
    r1[9]  = 1
    r1[10] = "phgenericap01"
    r1[11] = "10.1.0.11"
    r1[12] = "/usr/sap/TST"
    r1[13] = 100    # plain integer → size_gb=100, size_formula=None
    r1[14] = "xfs"
    r1[15] = "vg_sap"
    ws.append(r1)

    # Row 4: formula-string size row (MIN expression)
    r2 = blank_row()
    r2[0]  = 2
    r2[1]  = "PRD"
    r2[2]  = "TST"
    r2[3]  = "AP"
    r2[10] = "phgenericap01"
    r2[11] = "10.1.0.11"
    r2[12] = "/sapmnt"
    r2[13] = "MIN(256,1024)"   # formula string → size_gb=None, size_formula="MIN(256,1024)"
    r2[14] = "nfs"
    ws.append(r2)


def _build_securitygroup(wb):
    ws = wb.create_sheet("SecurityGroup")
    N = 17

    def blank_row():
        return [None] * N

    # Row 1: spacer
    ws.append(blank_row())

    # Row 2: top-level group labels
    ws.append(["Source", None, None, None, None, "Target", None, None, None,
               None, None, "ETC", None, "Work", None, None, None])

    # Row 3: column headers
    ws.append(["System", "Category", "Hostname", "IP", "Landscape",
               "System", "Category", "Hostname", "IP",
               "Port", "Protocol", "Expiration", "Purpose",
               "CUS FW", "CUS SG", "PSM FW", "PSM SG"])

    # Row 4: first data row
    r1 = blank_row()
    r1[0]  = "TST"
    r1[1]  = "AP"
    r1[2]  = "vhgenericap01"
    r1[3]  = "10.0.0.11"
    r1[4]  = "PRD"
    r1[5]  = "TST"
    r1[6]  = "DB"
    r1[7]  = "vhgenericdb01"
    r1[8]  = "10.0.0.21"
    r1[9]  = 30013          # integer port → "30013"
    r1[10] = "TCP"
    r1[12] = "HANA SQL"
    r1[13] = "O"            # cus_fw → True
    r1[14] = "X"            # cus_sg → False
    r1[15] = "O"            # psm_fw → True
    r1[16] = "X"            # psm_sg → False
    ws.append(r1)

    # Row 5: blank — must be SKIPPED (not terminate)
    ws.append(blank_row())

    # Row 6: second data row (blank row between should not stop parsing)
    r2 = blank_row()
    r2[0]  = "TST"
    r2[1]  = "APP"
    r2[2]  = "vhgenericap01"
    r2[3]  = "10.0.0.11"
    r2[5]  = "EXT"
    r2[8]  = "0.0.0.0/0"
    r2[9]  = "1128,1129"    # comma-separated port
    r2[10] = "TCP"
    r2[12] = "SAP Message Server"
    r2[13] = "O"
    r2[15] = "O"
    ws.append(r2)


def _build_sample_workbook(path):
    """Create a synthetic S2D workbook at *path* using generic placeholder data."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default "Sheet"
    _build_accinfo(wb)
    _build_nwinfo(wb)
    _build_serverinfo(wb)
    _build_filesystem(wb)
    _build_securitygroup(wb)
    wb.save(path)


# ---------------------------------------------------------------------------
# Session fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="session")
def sample_s2d_path(tmp_path_factory):
    path = tmp_path_factory.mktemp("s2d") / "sample_s2d.xlsx"
    _build_sample_workbook(path)
    return path


@pytest.fixture(scope="session")
def parsed(sample_s2d_path):
    return parse_s2d(str(sample_s2d_path))


# ---------------------------------------------------------------------------
# Top-level structure
# ---------------------------------------------------------------------------

class TestTopLevelStructure:
    def test_is_dict(self, parsed):
        assert isinstance(parsed, dict)

    def test_has_all_required_keys(self, parsed):
        for key in ("project_metadata", "global_network_context", "network",
                    "vmware_vm", "filesystem", "firewall_rule_candidates",
                    "parser_warnings", "ignored_labels"):
            assert key in parsed, f"missing key: {key}"


# ---------------------------------------------------------------------------
# Accinfo → project_metadata
# ---------------------------------------------------------------------------

class TestProjectMetadata:
    def test_customer_name(self, parsed):
        assert parsed["project_metadata"]["customer_name"] == "Generic Corp"

    def test_customer_id(self, parsed):
        assert parsed["project_metadata"]["customer_id"] == "GCI"

    def test_domain_name(self, parsed):
        assert parsed["project_metadata"]["domain_name"] == "sap.generic.net"

    def test_timezone(self, parsed):
        assert parsed["project_metadata"]["timezone"] == "KST (UTC+09:00)"


# ---------------------------------------------------------------------------
# Accinfo → global_network_context
# ---------------------------------------------------------------------------

class TestGlobalNetworkContext:
    def test_ip_range(self, parsed):
        assert parsed["global_network_context"]["ip_range"] == "10.0.0.0/22"

    def test_dr_ip_range(self, parsed):
        assert parsed["global_network_context"]["dr_ip_range"] == "10.4.0.0/22"

    def test_connectivity_type(self, parsed):
        assert parsed["global_network_context"]["connectivity_type"] == "Cloud Peering"

    def test_dns_primary(self, parsed):
        assert parsed["global_network_context"]["dns_primary"] == "10.0.0.10"

    def test_dns_secondary(self, parsed):
        assert parsed["global_network_context"]["dns_secondary"] == "10.0.0.11"

    def test_customer_networks_split(self, parsed):
        assert parsed["global_network_context"]["customer_networks"] == [
            "10.100.0.0/24", "10.100.1.0/24"
        ]


# ---------------------------------------------------------------------------
# Accinfo → ignored_labels & parser_warnings
# ---------------------------------------------------------------------------

class TestAccinfoClassification:
    def test_vpc_name_in_ignored_labels(self, parsed):
        assert any("vpc name" in lbl.lower() for lbl in parsed["ignored_labels"])

    def test_vpc_name_not_in_parser_warnings(self, parsed):
        assert not any("vpc name" in w.lower() for w in parsed["parser_warnings"])

    def test_no_parser_warnings_from_well_formed_data(self, parsed):
        # All labels in the synthetic sheet are known — no unexpected warnings
        assert parsed["parser_warnings"] == []


# ---------------------------------------------------------------------------
# NWInfo → network.subnets
# ---------------------------------------------------------------------------

class TestNetworkSubnets:
    def test_subnet_list_is_list(self, parsed):
        assert isinstance(parsed["network"]["subnets"], list)

    def test_total_subnet_count(self, parsed):
        # 3 customer-prd + 1 customer-dr + 1 internal-prd = 5
        assert len(parsed["network"]["subnets"]) == 5

    def test_customer_group_present(self, parsed):
        groups = {s["group"] for s in parsed["network"]["subnets"]}
        assert "customer" in groups

    def test_internal_group_present(self, parsed):
        groups = {s["group"] for s in parsed["network"]["subnets"]}
        assert "internal" in groups

    def test_prd_env_present(self, parsed):
        envs = {s["env"] for s in parsed["network"]["subnets"]}
        assert "prd" in envs

    def test_dr_env_present(self, parsed):
        envs = {s["env"] for s in parsed["network"]["subnets"]}
        assert "dr" in envs

    def test_purpose_key_admin(self, parsed):
        keys = {s["purpose_key"] for s in parsed["network"]["subnets"]}
        assert "admin" in keys

    def test_purpose_key_production_service(self, parsed):
        keys = {s["purpose_key"] for s in parsed["network"]["subnets"]}
        assert "production_service" in keys

    def test_purpose_key_storage(self, parsed):
        keys = {s["purpose_key"] for s in parsed["network"]["subnets"]}
        assert "storage" in keys

    def test_blank_row_does_not_terminate_nwinfo_parsing(self, parsed):
        # The NWInfo sheet has a blank row inside the customer-prd block.
        # All 3 customer-prd subnets must be present.
        customer_prd = [
            s for s in parsed["network"]["subnets"]
            if s["group"] == "customer" and s["env"] == "prd"
        ]
        assert len(customer_prd) == 3

    def test_ip_range_field_preserved(self, parsed):
        admin_subnets = [
            s for s in parsed["network"]["subnets"]
            if s["purpose_key"] == "admin" and s["env"] == "prd" and s["group"] == "customer"
        ]
        assert any(s["ip_range"] == "10.1.0.0/24" for s in admin_subnets)

    def test_hostname_pattern_preserved(self, parsed):
        srv_subnets = [
            s for s in parsed["network"]["subnets"]
            if s["purpose_key"] == "production_service"
        ]
        assert any(s["hostname_pattern"] == "vhgeneric*" for s in srv_subnets)


# ---------------------------------------------------------------------------
# NWInfo → network.routing_table
# ---------------------------------------------------------------------------

class TestRoutingTable:
    def test_routing_table_count(self, parsed):
        assert len(parsed["network"]["routing_table"]) == 2

    def test_routing_name(self, parsed):
        names = {r["routing_name"] for r in parsed["network"]["routing_table"]}
        assert "Customer Route" in names

    def test_target_cidr(self, parsed):
        route = next(r for r in parsed["network"]["routing_table"]
                     if r["routing_name"] == "Customer Route")
        assert route["target_cidr"] == "10.100.0.0/16"

    def test_source_ip(self, parsed):
        route = next(r for r in parsed["network"]["routing_table"]
                     if r["routing_name"] == "Customer Route")
        assert route["source_ip"] == "10.0.0.1"

    def test_remark_none(self, parsed):
        route = next(r for r in parsed["network"]["routing_table"]
                     if r["routing_name"] == "DR Route")
        assert route["remark"] is None


# ---------------------------------------------------------------------------
# ServerInfo → vmware_vm
# ---------------------------------------------------------------------------

class TestVmwareVm:
    def test_only_server_entries_included(self, parsed):
        # Sheet has 1 server + 1 vip; only server must appear
        assert len(parsed["vmware_vm"]) == 1

    def test_vhost(self, parsed):
        assert parsed["vmware_vm"][0]["identity"]["vhost"] == "vhgenericap01"

    def test_phost_stripped_of_concatenated_ip(self, parsed):
        # phost_raw = "phgenericap0110.1.0.11", admin_ip = "10.1.0.11"
        assert parsed["vmware_vm"][0]["identity"]["phost"] == "phgenericap01"

    def test_landscape(self, parsed):
        assert parsed["vmware_vm"][0]["identity"]["landscape"] == "PRD"

    def test_sid(self, parsed):
        assert parsed["vmware_vm"][0]["identity"]["sid"] == "TST"

    def test_os_version(self, parsed):
        assert parsed["vmware_vm"][0]["compute"]["os_version"] == "RHEL 9.4"

    def test_cpu_vcores(self, parsed):
        assert parsed["vmware_vm"][0]["compute"]["cpu_vcores"] == 8

    def test_memory_gb(self, parsed):
        assert parsed["vmware_vm"][0]["compute"]["memory_gb"] == 32

    def test_nfs_na_becomes_none(self, parsed):
        assert parsed["vmware_vm"][0]["compute"]["nfs_storage_gb"] is None

    def test_admin_ip(self, parsed):
        assert parsed["vmware_vm"][0]["network"]["admin_ip"] == "10.1.0.11"

    def test_sla(self, parsed):
        assert parsed["vmware_vm"][0]["availability"]["sla"] == "99.90%"

    def test_ha_true(self, parsed):
        assert parsed["vmware_vm"][0]["availability"]["ha"] is True

    def test_dr_false(self, parsed):
        assert parsed["vmware_vm"][0]["availability"]["dr"] is False

    def test_pacemaker_true(self, parsed):
        assert parsed["vmware_vm"][0]["availability"]["pacemaker"] is True

    def test_backup_enabled_true(self, parsed):
        assert parsed["vmware_vm"][0]["availability"]["backup_enabled"] is True

    def test_vm_or_bm(self, parsed):
        assert parsed["vmware_vm"][0]["metadata"]["vm_or_bm"] == "VM"


# ---------------------------------------------------------------------------
# FileSystem → filesystem
# ---------------------------------------------------------------------------

class TestFilesystem:
    def test_filesystem_count(self, parsed):
        assert len(parsed["filesystem"]) == 2

    def test_plain_integer_size(self, parsed):
        r = next(e for e in parsed["filesystem"] if e["mount_point"] == "/usr/sap/TST")
        assert r["size_gb"] == 100
        assert r["size_formula"] is None

    def test_formula_string_size(self, parsed):
        r = next(e for e in parsed["filesystem"] if e["mount_point"] == "/sapmnt")
        assert r["size_gb"] is None
        assert r["size_formula"] == "MIN(256,1024)"

    def test_ha_flag_normalized(self, parsed):
        r = next(e for e in parsed["filesystem"] if e["mount_point"] == "/usr/sap/TST")
        assert r["ha_flag"] is True

    def test_hostname_field(self, parsed):
        hostnames = {e["hostname"] for e in parsed["filesystem"]}
        assert "phgenericap01" in hostnames


# ---------------------------------------------------------------------------
# SecurityGroup → firewall_rule_candidates
# ---------------------------------------------------------------------------

class TestFirewallRuleCandidates:
    def test_blank_row_between_data_rows_skipped_not_terminated(self, parsed):
        # Sheet has: data / blank / data — both data rows must appear
        assert len(parsed["firewall_rule_candidates"]) == 2

    def test_integer_port_normalized_to_string(self, parsed):
        rule = next(r for r in parsed["firewall_rule_candidates"]
                    if r["purpose"] == "HANA SQL")
        assert rule["port"] == "30013"

    def test_comma_separated_port_preserved(self, parsed):
        rule = next(r for r in parsed["firewall_rule_candidates"]
                    if r["purpose"] == "SAP Message Server")
        assert rule["port"] == "1128,1129"

    def test_cus_fw_true(self, parsed):
        rule = next(r for r in parsed["firewall_rule_candidates"]
                    if r["purpose"] == "HANA SQL")
        assert rule["cus_fw"] is True

    def test_cus_sg_false(self, parsed):
        rule = next(r for r in parsed["firewall_rule_candidates"]
                    if r["purpose"] == "HANA SQL")
        assert rule["cus_sg"] is False

    def test_psm_fw_true(self, parsed):
        rule = next(r for r in parsed["firewall_rule_candidates"]
                    if r["purpose"] == "HANA SQL")
        assert rule["psm_fw"] is True

    def test_source_system(self, parsed):
        rule = next(r for r in parsed["firewall_rule_candidates"]
                    if r["purpose"] == "HANA SQL")
        assert rule["source_system"] == "TST"

    def test_protocol(self, parsed):
        rule = next(r for r in parsed["firewall_rule_candidates"]
                    if r["purpose"] == "HANA SQL")
        assert rule["protocol"] == "TCP"
