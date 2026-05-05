"""Debug script: parse NWInfo sheet from the sample workbook and write JSON.

Usage (from repo root):
    python scripts/debug_nwinfo.py [path/to/workbook.xlsx]

Defaults to the first .xlsx found in samples/ if no path is given.
Output is written to outputs/debug_nwinfo.json (pretty-printed).
"""
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from src.parser.nwinfo import parse_nwinfo

_KNOWN_PURPOSE_KEYS = {
    "production_service",
    "non_production_service",
    "admin",
    "heartbeat",
    "public",
    "storage",
    "backup",
}


def _find_sample() -> Path:
    samples = sorted((_REPO_ROOT / "samples").glob("*.xlsx"))
    if not samples:
        raise FileNotFoundError(
            "No .xlsx file found in samples/. "
            "Place the S2D workbook there and retry."
        )
    return samples[0]


def _summary(network: dict) -> None:
    subnets = network["subnets"]
    routing = network["routing_table"]

    group_counts = Counter(s["group"] for s in subnets)
    env_counts = Counter(s["env"] for s in subnets)
    unknown_purpose = [
        s for s in subnets if s["purpose_key"] not in _KNOWN_PURPOSE_KEYS
    ]

    print(f"  Subnets parsed:           {len(subnets)}")
    print(f"  Routing table entries:    {len(routing)}")
    print(f"  By group:")
    for group, count in sorted(group_counts.items()):
        print(f"    {group}: {count}")
    print(f"  By env:")
    for env, count in sorted(env_counts.items()):
        print(f"    {env}: {count}")
    print(f"  Unknown purpose_key:      {len(unknown_purpose)}")
    if unknown_purpose:
        for s in unknown_purpose:
            print(f"    purpose_raw={s['purpose_raw']!r}  purpose_key={s['purpose_key']!r}")


def main(workbook_path: Path) -> None:
    print(f"Opening: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    if "NWInfo" not in wb.sheetnames:
        raise KeyError(f"Sheet 'NWInfo' not found. Available: {wb.sheetnames}")

    ws = wb["NWInfo"]
    network, warnings = parse_nwinfo(ws)

    result = {
        "network":         network,
        "parser_warnings": warnings,
    }

    out_path = _REPO_ROOT / "outputs" / "debug_nwinfo.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    print(f"Written:  {out_path}")
    _summary(network)
    if warnings:
        print(f"  Warnings ({len(warnings)}):")
        for w in warnings:
            print(f"    {w}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else _find_sample()
    main(path)
