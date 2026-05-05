"""Debug script: parse SecurityGroup sheet from the sample workbook and write JSON.

Usage (from repo root):
    python scripts/debug_securitygroup.py [path/to/workbook.xlsx]

Defaults to the first .xlsx found in samples/ if no path is given.
Output is written to outputs/debug_securitygroup.json (pretty-printed).
"""
import json
import sys
from pathlib import Path

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from src.parser.securitygroup import parse_securitygroup


def _find_sample() -> Path:
    samples = sorted((_REPO_ROOT / "samples").glob("*.xlsx"))
    if not samples:
        raise FileNotFoundError(
            "No .xlsx file found in samples/. "
            "Place the S2D workbook there and retry."
        )
    return samples[0]


def _count_skipped(ws) -> int:
    """Count rows that were skipped (cols 0-12 all None), excluding headers."""
    skipped = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        if all(row[j] is None for j in range(min(13, len(row)))):
            skipped += 1
    return skipped


def _summary(candidates: list[dict], skipped: int) -> None:
    total = len(candidates)
    single_port = sum(
        1 for c in candidates
        if c["port"] is not None and not c["port_expression"]
    )
    no_port = sum(1 for c in candidates if c["port"] is None)
    expressions = sum(1 for c in candidates if c["port_expression"])

    missing_source = sum(
        1 for c in candidates
        if not c["source_hostname"] and not c["source_ip"]
    )
    missing_target = sum(
        1 for c in candidates
        if not c["target_hostname"] and not c["target_ip"]
    )
    missing_either = sum(
        1 for c in candidates
        if (not c["source_hostname"] and not c["source_ip"])
        or (not c["target_hostname"] and not c["target_ip"])
    )

    print(f"  Rule candidates parsed:         {total}")
    print(f"  Rows skipped (blank cols 0-12): {skipped}")
    print(f"  Rules with single port:         {single_port}")
    print(f"  Rules with no port:             {no_port}")
    print(f"  Rules with port expression:     {expressions}")
    print(f"  Missing source identifier:      {missing_source}")
    print(f"  Missing target identifier:      {missing_target}")
    print(f"  Missing source or target:       {missing_either}")


def main(workbook_path: Path) -> None:
    print(f"Opening: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    if "SecurityGroup" not in wb.sheetnames:
        raise KeyError(f"Sheet 'SecurityGroup' not found. Available: {wb.sheetnames}")

    ws = wb["SecurityGroup"]
    skipped = _count_skipped(ws)
    candidates, warnings = parse_securitygroup(ws)

    result = {
        "firewall_rule_candidates": candidates,
        "parser_warnings":          warnings,
    }

    out_path = _REPO_ROOT / "outputs" / "debug_securitygroup.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    print(f"Written:  {out_path}")
    _summary(candidates, skipped)
    if warnings:
        print(f"  Warnings ({len(warnings)}):")
        for w in warnings:
            print(f"    {w}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else _find_sample()
    main(path)
