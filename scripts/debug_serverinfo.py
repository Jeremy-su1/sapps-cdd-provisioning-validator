"""Debug script: parse ServerInfo sheet from the sample workbook and write JSON.

Usage (from repo root):
    python scripts/debug_serverinfo.py [path/to/workbook.xlsx]

Defaults to the first .xlsx found in samples/ if no path is given.
Output is written to outputs/debug_serverinfo.json (pretty-printed).
"""
import json
import sys
from pathlib import Path

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from src.parser.serverinfo import parse_serverinfo


def _find_sample() -> Path:
    samples = sorted((_REPO_ROOT / "samples").glob("*.xlsx"))
    if not samples:
        raise FileNotFoundError(
            "No .xlsx file found in samples/. "
            "Place the S2D workbook there and retry."
        )
    return samples[0]


def main(workbook_path: Path) -> None:
    print(f"Opening: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    if "ServerInfo" not in wb.sheetnames:
        raise KeyError(f"Sheet 'ServerInfo' not found. Available: {wb.sheetnames}")

    ws = wb["ServerInfo"]
    vmware_vm, warnings = parse_serverinfo(ws)

    result = {
        "vmware_vm":       vmware_vm,
        "parser_warnings": warnings,
    }

    out_path = _REPO_ROOT / "outputs" / "debug_serverinfo.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    print(f"Written:  {out_path}")
    print(f"  VMs parsed: {len(vmware_vm)}")
    if warnings:
        print(f"  ⚠  {len(warnings)} parser warning(s):")
        for w in warnings:
            print(f"     {w}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else _find_sample()
    main(path)
