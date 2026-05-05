"""Debug script: parse FileSystem sheet from the sample workbook and write JSON.

Usage (from repo root):
    python scripts/debug_filesystem.py [path/to/workbook.xlsx]

Defaults to the first .xlsx found in samples/ if no path is given.
Output is written to outputs/debug_filesystem.json (pretty-printed).
"""
import json
import sys
from pathlib import Path

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from src.parser.filesystem import parse_filesystem


def _find_sample() -> Path:
    samples = sorted((_REPO_ROOT / "samples").glob("*.xlsx"))
    if not samples:
        raise FileNotFoundError(
            "No .xlsx file found in samples/. "
            "Place the S2D workbook there and retry."
        )
    return samples[0]


def _summary(entries: list[dict]) -> None:
    total = len(entries)
    with_size_gb = sum(1 for e in entries if e["size_gb"] is not None)
    with_formula = sum(1 for e in entries if e["size_formula"] is not None)
    missing_hostname = sum(1 for e in entries if not e["hostname"])
    missing_admin_ip = sum(1 for e in entries if not e["admin_ip"])
    missing_join = sum(
        1 for e in entries if not e["hostname"] or not e["admin_ip"]
    )

    print(f"  Entries parsed:          {total}")
    print(f"  With numeric size_gb:    {with_size_gb}")
    print(f"  With size_formula:       {with_formula}")
    print(f"  Missing hostname:        {missing_hostname}")
    print(f"  Missing admin_ip:        {missing_admin_ip}")
    print(f"  Missing hostname or ip:  {missing_join}")


def main(workbook_path: Path) -> None:
    print(f"Opening: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    if "FileSystem" not in wb.sheetnames:
        raise KeyError(f"Sheet 'FileSystem' not found. Available: {wb.sheetnames}")

    ws = wb["FileSystem"]
    entries, warnings = parse_filesystem(ws)

    result = {
        "filesystem":      entries,
        "parser_warnings": warnings,
    }

    out_path = _REPO_ROOT / "outputs" / "debug_filesystem.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    print(f"Written:  {out_path}")
    _summary(entries)
    if warnings:
        print(f"  Warnings ({len(warnings)}):")
        for w in warnings:
            print(f"    {w}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else _find_sample()
    main(path)
