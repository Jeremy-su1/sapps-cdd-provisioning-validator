"""Debug script: parse Accinfo sheet from the sample workbook and write JSON.

Usage (from repo root):
    python scripts/debug_accinfo.py [path/to/workbook.xlsx]

Defaults to the first .xlsx found in samples/ if no path is given.
Output is written to outputs/debug_accinfo.json (pretty-printed).
"""
import json
import sys
from pathlib import Path

import openpyxl

# Allow running from any working directory
_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))

from src.parser.accinfo import parse_accinfo


def _find_sample() -> Path:
    samples = sorted((_REPO_ROOT / "samples").glob("*.xlsx"))
    if not samples:
        raise FileNotFoundError(
            "No .xlsx file found in samples/. "
            "Place the S2D workbook there and retry."
        )
    return samples[0]


def main(workbook_path: Path) -> None:
    print(f"Opening: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    if "Accinfo" not in wb.sheetnames:
        raise KeyError(f"Sheet 'Accinfo' not found. Available: {wb.sheetnames}")

    ws = wb["Accinfo"]
    project_metadata, network_context, warnings, ignored_labels = parse_accinfo(ws)

    result = {
        "project_metadata":       project_metadata,
        "global_network_context": network_context,
        "parser_warnings":        warnings,
        "ignored_labels":         ignored_labels,
    }

    out_path = _REPO_ROOT / "outputs" / "debug_accinfo.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2))

    print(f"Written: {out_path}")
    if warnings:
        print(f"  ⚠  {len(warnings)} parser warning(s):")
        for w in warnings:
            print(f"     {w}")
    if ignored_labels:
        print(f"  ℹ  {len(ignored_labels)} ignored label(s): {ignored_labels}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else _find_sample()
    main(path)
